# -*- coding: utf-8 -*-
"""
Buscador de Arquivos — CNPJ/Folha e Holerites
===============================================
Programa com duas abas:
  Aba 1 - Busca de Folhas por CNPJ (lógica original)
  Aba 2 - Busca de Holerites por Posto (lógica nova)
"""

import os
import re
import shutil
import threading
import traceback
from datetime import datetime

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

try:
    import openpyxl
except ImportError:
    openpyxl = None


# ======================================================================
# UTILITÁRIOS COMUNS
# ======================================================================

def _exigir_openpyxl():
    if openpyxl is None:
        raise RuntimeError(
            "A biblioteca 'openpyxl' não está instalada.\n"
            "Abra o Prompt de Comando e rode: pip install openpyxl"
        )

def _salvar_txt(caminho, titulo, linhas_texto):
    """Salva um arquivo .txt com cabeçalho de data/hora."""
    with open(caminho, "w", encoding="utf-8") as f:
        f.write(f"{titulo} — gerado em {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
        f.write("=" * 60 + "\n\n")
        for linha in linhas_texto:
            f.write(linha + "\n")


# ======================================================================
# ABA 1 — BUSCA DE FOLHAS POR CNPJ
# ======================================================================

def limpar_documento(valor):
    """Remove tudo que não for dígito de um CNPJ/CPF."""
    if valor is None:
        return ""
    return re.sub(r"\D", "", str(valor))


def ler_planilha_cnpj(caminho_excel):
    """Lê planilha com colunas TOMADOR e CNPJ."""
    _exigir_openpyxl()
    wb = openpyxl.load_workbook(caminho_excel, data_only=True)
    sheet = wb.active
    linhas = list(sheet.iter_rows(values_only=True))
    if not linhas:
        raise ValueError("A planilha está vazia.")

    cab = [str(c).strip().upper() if c is not None else "" for c in linhas[0]]
    try:
        idx_t = cab.index("TOMADOR")
    except ValueError:
        raise ValueError("Coluna 'TOMADOR' não encontrada na planilha.")
    try:
        idx_c = cab.index("CNPJ")
    except ValueError:
        raise ValueError("Coluna 'CNPJ' não encontrada na planilha.")

    resultado = []
    for linha in linhas[1:]:
        if not linha or (linha[idx_t] is None and linha[idx_c] is None):
            continue
        tomador = str(linha[idx_t]).strip() if linha[idx_t] is not None else ""
        cnpj_orig = str(linha[idx_c]).strip() if linha[idx_c] is not None else ""
        cnpj_limpo = limpar_documento(linha[idx_c])
        if not tomador and not cnpj_limpo:
            continue
        resultado.append((tomador, cnpj_limpo, cnpj_orig))
    return resultado


def encontrar_pastas_do_cnpj(pasta_raiz, cnpj_limpo):
    todos = []
    for atual, subpastas, _ in os.walk(pasta_raiz):
        for nome in subpastas:
            if limpar_documento(nome) == cnpj_limpo:
                todos.append(os.path.join(atual, nome))
    todos.sort(key=len)
    de_topo = []
    for c in todos:
        cn = os.path.normpath(c)
        if not any(cn.startswith(os.path.normpath(p) + os.sep) for p in de_topo):
            de_topo.append(c)
    return de_topo


def encontrar_subpasta_repetida(pasta_cnpj, cnpj_limpo):
    candidatas = []
    for atual, subpastas, _ in os.walk(pasta_cnpj):
        for nome in subpastas:
            if limpar_documento(nome) == cnpj_limpo:
                candidatas.append(os.path.join(atual, nome))
    if not candidatas:
        return None
    candidatas.sort(key=len)
    return candidatas[0]


def encontrar_arquivo_folha(pasta_cnpj, cnpj_limpo, log_callback=None):
    alvo = "folha"
    subpasta = encontrar_subpasta_repetida(pasta_cnpj, cnpj_limpo)
    if subpasta:
        for atual, _, arquivos in os.walk(subpasta):
            for nome in arquivos:
                if alvo in os.path.splitext(nome)[0].lower():
                    return os.path.join(atual, nome)
        return None
    if log_callback:
        log_callback(f"    Aviso: subpasta repetida não encontrada em '{pasta_cnpj}'. "
                     f"Procurando em toda a pasta.")
    for atual, _, arquivos in os.walk(pasta_cnpj):
        for nome in arquivos:
            if alvo in os.path.splitext(nome)[0].lower():
                return os.path.join(atual, nome)
    return None


def processar_cnpj(caminho_excel, pasta_busca, pasta_destino, log_callback, progresso_callback):
    os.makedirs(pasta_destino, exist_ok=True)
    log_callback("Lendo planilha...")
    linhas = ler_planilha_cnpj(caminho_excel)
    total = len(linhas)
    log_callback(f"{total} linha(s) encontrada(s).\n")

    try:
        pastas_raiz = [n for n in sorted(os.listdir(pasta_busca))
                       if os.path.isdir(os.path.join(pasta_busca, n))]
        log_callback(f"Pasta de busca: {pasta_busca}")
        log_callback(f"Subpastas na raiz ({len(pastas_raiz)}): "
                     f"{', '.join(pastas_raiz[:30])}"
                     f"{'...' if len(pastas_raiz) > 30 else ''}\n")
    except Exception as e:
        log_callback(f"Aviso: não foi possível listar a pasta de busca ({e}).\n")

    nao_encontrados, duplicados, erros = [], [], []
    copiados = 0

    for i, (tomador, cnpj_limpo, cnpj_orig) in enumerate(linhas, 1):
        progresso_callback(i, total)
        if not cnpj_limpo:
            log_callback(f"[{i}/{total}] Tomador '{tomador}': CNPJ vazio. Pulando.")
            nao_encontrados.append(f"TOMADOR: {tomador}\tCNPJ: {cnpj_orig or '(vazio)'}")
            continue

        log_callback(f"[{i}/{total}] Procurando CNPJ '{cnpj_orig}' -> '{cnpj_limpo}'...")
        pastas = encontrar_pastas_do_cnpj(pasta_busca, cnpj_limpo)

        if not pastas:
            log_callback(f"[{i}/{total}] Tomador '{tomador}': pasta não encontrada.")
            nao_encontrados.append(f"TOMADOR: {tomador}\tCNPJ: {cnpj_orig}")
            continue

        if len(pastas) > 1:
            log_callback(f"[{i}/{total}] Tomador '{tomador}': {len(pastas)} pastas encontradas, usando a primeira.")
            duplicados.append(f"TOMADOR: {tomador}\tCNPJ: {cnpj_orig}\n"
                              + "\n".join(f"    - {p}" for p in pastas))

        arquivo_orig = encontrar_arquivo_folha(pastas[0], cnpj_limpo, log_callback)
        if not arquivo_orig:
            log_callback(f"[{i}/{total}] Tomador '{tomador}': pasta encontrada mas sem arquivo 'FOLHA'.")
            nao_encontrados.append(f"TOMADOR: {tomador}\tCNPJ: {cnpj_orig}")
            continue

        _, ext = os.path.splitext(arquivo_orig)
        nome_seg = re.sub(r'[\\/:*?"<>|]', "_", str(tomador)) or "SEM_TOMADOR"
        destino_final = os.path.join(pasta_destino, f"{nome_seg}{ext}")
        if os.path.exists(destino_final):
            base, e2 = os.path.splitext(f"{nome_seg}{ext}")
            k = 2
            while os.path.exists(destino_final):
                destino_final = os.path.join(pasta_destino, f"{base}_{k}{e2}")
                k += 1

        try:
            shutil.copy2(arquivo_orig, destino_final)
            copiados += 1
            log_callback(f"[{i}/{total}] Tomador '{tomador}': OK -> {os.path.basename(destino_final)}")
        except Exception as e:
            erros.append(f"TOMADOR: {tomador}\tCNPJ: {cnpj_orig}\tERRO: {e}")
            log_callback(f"[{i}/{total}] Tomador '{tomador}': ERRO ao copiar — {e}")

    if nao_encontrados:
        p = os.path.join(pasta_destino, "nao_encontrados.txt")
        _salvar_txt(p, "Tomadores/CNPJs não encontrados", nao_encontrados)
        log_callback(f"\n'nao_encontrados.txt' gerado com {len(nao_encontrados)} item(ns).")
    if duplicados:
        p = os.path.join(pasta_destino, "duplicados.txt")
        _salvar_txt(p, "CNPJs com mais de uma pasta (usada a primeira)", duplicados)
        log_callback(f"'duplicados.txt' gerado com {len(duplicados)} caso(s).")
    if erros:
        p = os.path.join(pasta_destino, "erros_copia.txt")
        _salvar_txt(p, "Erros ao copiar", erros)
        log_callback(f"'erros_copia.txt' gerado com {len(erros)} erro(s).")

    return {"total": total, "copiados": copiados,
            "nao_encontrados": len(nao_encontrados),
            "duplicados": len(duplicados), "erros": len(erros)}


# ======================================================================
# ABA 2 — BUSCA DE HOLERITES POR POSTO
# ======================================================================

def ler_planilha_posto(caminho_excel):
    """Lê planilha com coluna POSTO. Retorna lista de strings com o número do posto."""
    _exigir_openpyxl()
    wb = openpyxl.load_workbook(caminho_excel, data_only=True)
    sheet = wb.active
    linhas = list(sheet.iter_rows(values_only=True))
    if not linhas:
        raise ValueError("A planilha está vazia.")

    cab = [str(c).strip().upper() if c is not None else "" for c in linhas[0]]
    try:
        idx_p = cab.index("TOMADOR")
    except ValueError:
        raise ValueError("Coluna 'TOMADOR' não encontrada na planilha.")

    resultado = []
    for linha in linhas[1:]:
        if not linha or linha[idx_p] is None:
            continue
        posto = str(linha[idx_p]).strip()
        if posto:
            resultado.append(posto)
    return resultado


def encontrar_pasta_do_posto(pasta_raiz, numero_posto):
    """
    Procura, em qualquer nível dentro de pasta_raiz, uma pasta cujo nome
    contenha o número do posto. Retorna a lista de pastas encontradas.
    Usa correspondência exata do número dentro do nome da pasta para evitar
    falsos positivos (ex: posto '10' não deve bater com '100').
    """
    encontradas = []
    for atual, subpastas, _ in os.walk(pasta_raiz):
        for nome in subpastas:
            # Verifica se o número do posto aparece como palavra/token isolado
            # dentro do nome da pasta (ex: "POSTO 42", "42 - Nome", "042")
            nome_norm = re.sub(r"\D+", " ", nome).strip()
            tokens = nome_norm.split()
            # Também compara removendo zeros à esquerda para cobrir "042" == "42"
            if any(t == numero_posto or t.lstrip("0") == numero_posto.lstrip("0")
                   for t in tokens):
                encontradas.append(os.path.join(atual, nome))
    # Filtra apenas os de topo (sem aninhamento)
    encontradas.sort(key=len)
    de_topo = []
    for c in encontradas:
        cn = os.path.normpath(c)
        if not any(cn.startswith(os.path.normpath(p) + os.sep) for p in de_topo):
            de_topo.append(c)
    return de_topo


def encontrar_holerites_do_posto(pasta_posto):
    """
    Dentro da pasta do posto (sem entrar em subpastas), procura arquivos
    que contenham 'holerite' E ('seg' ou 'ser') no nome (sem diferenciar
    maiúsculas/minúsculas).
    Retorna dicionário: {'SEG': [caminhos], 'SER': [caminhos]}
    """
    resultado = {"SEG": [], "SER": []}
    try:
        arquivos = [f for f in os.listdir(pasta_posto)
                    if os.path.isfile(os.path.join(pasta_posto, f))]
    except Exception:
        return resultado

    for nome in arquivos:
        nome_lower = os.path.splitext(nome)[0].lower()
        tem_holerite = "holerite" in nome_lower
        tem_seg = "seg" in nome_lower
        tem_ser = "ser" in nome_lower
        if tem_holerite and tem_seg:
            resultado["SEG"].append(os.path.join(pasta_posto, nome))
        elif tem_holerite and tem_ser:
            resultado["SER"].append(os.path.join(pasta_posto, nome))
    return resultado


def processar_holerites(caminho_excel, pasta_busca, pasta_destino, log_callback, progresso_callback):
    pasta_seg = os.path.join(pasta_destino, "Holerites SEG")
    pasta_ser = os.path.join(pasta_destino, "Holerites SER")
    os.makedirs(pasta_seg, exist_ok=True)
    os.makedirs(pasta_ser, exist_ok=True)

    log_callback("Lendo planilha de postos...")
    postos = ler_planilha_posto(caminho_excel)
    total = len(postos)
    log_callback(f"{total} posto(s) encontrado(s).\n")

    try:
        pastas_raiz = [n for n in sorted(os.listdir(pasta_busca))
                       if os.path.isdir(os.path.join(pasta_busca, n))]
        log_callback(f"Pasta de busca: {pasta_busca}")
        log_callback(f"Subpastas na raiz ({len(pastas_raiz)}): "
                     f"{', '.join(pastas_raiz[:30])}"
                     f"{'...' if len(pastas_raiz) > 30 else ''}\n")
    except Exception as e:
        log_callback(f"Aviso: não foi possível listar a pasta de busca ({e}).\n")

    nao_encontrados, erros = [], []
    copiados_seg = copiados_ser = 0

    for i, posto in enumerate(postos, 1):
        progresso_callback(i, total)
        log_callback(f"[{i}/{total}] Posto '{posto}'...")

        pastas = encontrar_pasta_do_posto(pasta_busca, posto)
        if not pastas:
            log_callback(f"[{i}/{total}] Posto '{posto}': pasta não encontrada.")
            nao_encontrados.append(f"POSTO: {posto}")
            continue

        if len(pastas) > 1:
            log_callback(f"[{i}/{total}] Posto '{posto}': {len(pastas)} pastas encontradas, usando a primeira.")

        pasta_posto = pastas[0]
        holerites = encontrar_holerites_do_posto(pasta_posto)

        achou_algo = False
        for tipo, arquivos in holerites.items():
            pasta_tipo = pasta_seg if tipo == "SEG" else pasta_ser
            for origem in arquivos:
                nome_orig = os.path.basename(origem)
                destino_final = os.path.join(pasta_tipo, nome_orig)
                # Evita sobrescrita silenciosa
                if os.path.exists(destino_final):
                    base, ext = os.path.splitext(nome_orig)
                    k = 2
                    while os.path.exists(destino_final):
                        destino_final = os.path.join(pasta_tipo, f"{base}_{k}{ext}")
                        k += 1
                try:
                    shutil.copy2(origem, destino_final)
                    if tipo == "SEG":
                        copiados_seg += 1
                    else:
                        copiados_ser += 1
                    achou_algo = True
                    log_callback(f"    [{tipo}] {nome_orig} -> {os.path.basename(destino_final)}")
                except Exception as e:
                    erros.append(f"POSTO: {posto}\tARQUIVO: {nome_orig}\tERRO: {e}")
                    log_callback(f"    [{tipo}] ERRO ao copiar '{nome_orig}': {e}")

        if not achou_algo:
            log_callback(f"[{i}/{total}] Posto '{posto}': pasta encontrada mas sem holerites SEG/SER.")
            nao_encontrados.append(f"POSTO: {posto}\t(pasta encontrada, sem holerites SEG/SER)")

    if nao_encontrados:
        p = os.path.join(pasta_destino, "nao_encontrados_holerites.txt")
        _salvar_txt(p, "Postos sem holerites encontrados", nao_encontrados)
        log_callback(f"\n'nao_encontrados_holerites.txt' gerado com {len(nao_encontrados)} item(ns).")
    if erros:
        p = os.path.join(pasta_destino, "erros_copia_holerites.txt")
        _salvar_txt(p, "Erros ao copiar holerites", erros)
        log_callback(f"'erros_copia_holerites.txt' gerado com {len(erros)} erro(s).")

    return {"total": total,
            "copiados_seg": copiados_seg, "copiados_ser": copiados_ser,
            "nao_encontrados": len(nao_encontrados), "erros": len(erros)}


# ======================================================================
# INTERFACE GRÁFICA — COMPONENTE DE ABA GENÉRICO
# ======================================================================

class AbaBase:
    """Classe base com os widgets comuns às duas abas (seleção + log + progresso)."""

    def __init__(self, notebook, titulo_aba):
        self.frame = ttk.Frame(notebook)
        notebook.add(self.frame, text=titulo_aba)
        self._widgets_comuns()

    def _widgets_comuns(self):
        self._vars = {}
        self._frames_selecao = ttk.Frame(self.frame)
        self._frames_selecao.pack(fill="x", padx=10, pady=6)

        frame_acao = ttk.Frame(self.frame)
        frame_acao.pack(fill="x", padx=10, pady=4)
        self.botao_iniciar = ttk.Button(frame_acao, text="Iniciar", command=self._iniciar)
        self.botao_iniciar.pack(side="left")
        self.barra = ttk.Progressbar(frame_acao, mode="determinate")
        self.barra.pack(side="left", fill="x", expand=True, padx=10)
        self.label_status = ttk.Label(frame_acao, text="")
        self.label_status.pack(side="left")

        frame_log = ttk.LabelFrame(self.frame, text="Andamento")
        frame_log.pack(fill="both", expand=True, padx=10, pady=6)
        self.texto_log = tk.Text(frame_log, wrap="word", state="disabled")
        self.texto_log.pack(side="left", fill="both", expand=True)
        sb = ttk.Scrollbar(frame_log, command=self.texto_log.yview)
        sb.pack(side="right", fill="y")
        self.texto_log.configure(yscrollcommand=sb.set)

    def _linha_selecao(self, parent, label, key, comando):
        lf = ttk.LabelFrame(parent, text=label)
        lf.pack(fill="x", pady=3)
        var = tk.StringVar()
        self._vars[key] = var
        f = ttk.Frame(lf)
        f.pack(fill="x", padx=8, pady=5)
        ttk.Entry(f, textvariable=var).pack(side="left", fill="x", expand=True)
        ttk.Button(f, text="Escolher...", command=comando).pack(side="left", padx=6)

    def _escolher_arquivo(self, key, titulo):
        p = filedialog.askopenfilename(
            title=titulo,
            filetypes=[("Excel", "*.xlsx *.xlsm"), ("Todos os arquivos", "*.*")])
        if p:
            self._vars[key].set(p)

    def _escolher_pasta(self, key, titulo):
        p = filedialog.askdirectory(title=titulo)
        if p:
            self._vars[key].set(p)

    def _log(self, msg):
        self.texto_log.configure(state="normal")
        self.texto_log.insert("end", msg + "\n")
        self.texto_log.see("end")
        self.texto_log.configure(state="disabled")

    def _progresso(self, atual, total):
        self.barra["maximum"] = total
        self.barra["value"] = atual
        self.label_status.configure(text=f"{atual}/{total}")

    def _iniciar(self):
        raise NotImplementedError

    def _rodar_em_thread(self, fn, *args):
        self.botao_iniciar.configure(state="disabled")
        self.texto_log.configure(state="normal")
        self.texto_log.delete("1.0", "end")
        self.texto_log.configure(state="disabled")
        self.barra["value"] = 0
        root = self.frame.winfo_toplevel()

        def _run():
            try:
                resultado = fn(
                    *args,
                    log_callback=lambda m: root.after(0, self._log, m),
                    progresso_callback=lambda a, t: root.after(0, self._progresso, a, t),
                )
                root.after(0, self._finalizar_sucesso, resultado)
            except Exception as e:
                det = traceback.format_exc()
                root.after(0, self._finalizar_erro, str(e), det)

        threading.Thread(target=_run, daemon=True).start()

    def _finalizar_erro(self, msg, detalhe):
        self.botao_iniciar.configure(state="normal")
        self._log(f"\nERRO: {msg}")
        messagebox.showerror("Erro", f"Ocorreu um erro:\n\n{msg}")
        print(detalhe)

    def _finalizar_sucesso(self, resultado):
        raise NotImplementedError


# ======================================================================
# ABA 1 — INTERFACE
# ======================================================================

class AbaCNPJ(AbaBase):
    def __init__(self, notebook):
        super().__init__(notebook, "Folhas por CNPJ")
        self._linha_selecao(self._frames_selecao,
                            "1. Planilha Excel (colunas TOMADOR e CNPJ)",
                            "excel", lambda: self._escolher_arquivo("excel", "Selecione a planilha Excel"))
        self._linha_selecao(self._frames_selecao,
                            "2. Pasta onde estão as pastas dos CNPJs",
                            "busca", lambda: self._escolher_pasta("busca", "Selecione a pasta de busca"))
        self._linha_selecao(self._frames_selecao,
                            "3. Pasta de destino",
                            "destino", lambda: self._escolher_pasta("destino", "Selecione a pasta de destino"))

    def _iniciar(self):
        excel = self._vars["excel"].get().strip()
        busca = self._vars["busca"].get().strip()
        destino = self._vars["destino"].get().strip()
        if not excel or not os.path.isfile(excel):
            messagebox.showerror("Erro", "Selecione um arquivo Excel válido.")
            return
        if not busca or not os.path.isdir(busca):
            messagebox.showerror("Erro", "Selecione uma pasta de busca válida.")
            return
        if not destino:
            messagebox.showerror("Erro", "Selecione a pasta de destino.")
            return
        self._rodar_em_thread(processar_cnpj, excel, busca, destino)

    def _finalizar_sucesso(self, r):
        self.botao_iniciar.configure(state="normal")
        self._log(f"\nConcluído!\nTotal: {r['total']} | Copiados: {r['copiados']} | "
                  f"Não encontrados: {r['nao_encontrados']} | "
                  f"Duplicados: {r['duplicados']} | Erros: {r['erros']}")
        messagebox.showinfo("Concluído",
                            f"Processo finalizado.\n\n"
                            f"Copiados: {r['copiados']}\n"
                            f"Não encontrados: {r['nao_encontrados']}\n"
                            f"Duplicados: {r['duplicados']}\n"
                            f"Erros: {r['erros']}")


# ======================================================================
# ABA 2 — INTERFACE
# ======================================================================

class AbaHolerites(AbaBase):
    def __init__(self, notebook):
        super().__init__(notebook, "Holerites por Posto")
        self._linha_selecao(self._frames_selecao,
                            "1. Planilha Excel (mesma planilha com colunas TOMADOR e CNPJ)",
                            "excel", lambda: self._escolher_arquivo("excel", "Selecione a planilha Excel"))
        self._linha_selecao(self._frames_selecao,
                            "2. Pasta onde estão as pastas dos Postos",
                            "busca", lambda: self._escolher_pasta("busca", "Selecione a pasta de busca"))
        self._linha_selecao(self._frames_selecao,
                            "3. Pasta de destino (serão criadas 'Holerites SEG' e 'Holerites SER' aqui)",
                            "destino", lambda: self._escolher_pasta("destino", "Selecione a pasta de destino"))

    def _iniciar(self):
        excel = self._vars["excel"].get().strip()
        busca = self._vars["busca"].get().strip()
        destino = self._vars["destino"].get().strip()
        if not excel or not os.path.isfile(excel):
            messagebox.showerror("Erro", "Selecione um arquivo Excel válido.")
            return
        if not busca or not os.path.isdir(busca):
            messagebox.showerror("Erro", "Selecione uma pasta de busca válida.")
            return
        if not destino:
            messagebox.showerror("Erro", "Selecione a pasta de destino.")
            return
        self._rodar_em_thread(processar_holerites, excel, busca, destino)

    def _finalizar_sucesso(self, r):
        self.botao_iniciar.configure(state="normal")
        self._log(f"\nConcluído!\nTotal de postos: {r['total']} | "
                  f"SEG copiados: {r['copiados_seg']} | SER copiados: {r['copiados_ser']} | "
                  f"Não encontrados: {r['nao_encontrados']} | Erros: {r['erros']}")
        messagebox.showinfo("Concluído",
                            f"Processo finalizado.\n\n"
                            f"Holerites SEG copiados: {r['copiados_seg']}\n"
                            f"Holerites SER copiados: {r['copiados_ser']}\n"
                            f"Postos não encontrados: {r['nao_encontrados']}\n"
                            f"Erros: {r['erros']}")


# ======================================================================
# APLICAÇÃO PRINCIPAL
# ======================================================================

def main():
    root = tk.Tk()
    root.title("Buscador de Arquivos")
    root.geometry("760x600")
    root.minsize(700, 540)

    try:
        style = ttk.Style()
        if "vista" in style.theme_names():
            style.theme_use("vista")
    except Exception:
        pass

    notebook = ttk.Notebook(root)
    notebook.pack(fill="both", expand=True, padx=10, pady=10)

    AbaCNPJ(notebook)
    AbaHolerites(notebook)

    root.mainloop()


if __name__ == "__main__":
    main()
